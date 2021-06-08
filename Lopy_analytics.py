import json
import os
from math import ceil, isnan

import openpyxl
import pandas as pd

from Entities.Sigfox import Sigfox


def extract_data(file, output):
    pd.set_option('display.max_columns', None)

    current_experiment = int(file[file.rfind("_") + 1:file.rfind(".")])
    current_column = 5 + current_experiment

    packet_size = int(file[file.rfind("_", 0, file.rfind("_")) + 1:file.rfind("_")])
    fer = foldername[:foldername.find("_")]
    sheet = spreadsheet[f'Case {packet_size} FER {fer}']
    sheet["E1"] = packet_size

    header_bytes = 1 if packet_size <= 300 else 2
    profile = Sigfox("UPLINK", "ACK ON ERROR", header_bytes)

    fragments = ceil((packet_size * 8) / (profile.MTU - profile.HEADER_LENGTH))
    sheet["E2"] = fragments
    windows = ceil(fragments / profile.WINDOW_SIZE)
    sheet["E3"] = windows

    output.write(f"-------------- RESULTS FOR {file} --------------\n\n")

    with open(file, encoding='ISO-8859-1') as json_file:
        data = json.load(json_file)

    tx_json = {'fragmentation_time': data['fragmentation_time'],
               'payload_size': data['payload_size'],
               'total_number_of_fragments': data['total_number_of_fragments'],
               'total_transmission_time': data['total_transmission_time']}
    # output.write(f"tx_json =\n{tx_json}\n\n")
    df_tx = pd.DataFrame(tx_json, index=[0])
    df_tx = df_tx.T
    # output.write(f"dl_tx =\n{df_tx}\n\n")

    assert data['fragments']
    df1 = pd.read_json(str(json.dumps(data['fragments'], sort_keys=True)))
    # print(df1)
    # df1.astype({"Column 1": int, "Column 2": int, "Column 3": int})
    # df1.astype({"Column 6": bool})
    # print(df1.dtypes)
    df1_transposed = df1.T  # or df1.transpose()
    # output.write(f"df1_transposed =\n{df1_transposed} \n\n")
    df1_transposed.astype(
        {"downlink_enable": bool, "sending_start": float, "ack_received": bool, "data": str, "sending_end": float,
         "FCN": str, "fragment_size": int, "ack": str, "timeout": int, "RULE_ID": str, "rssi": int, "W": str,
         "send_time": float})
    # print(df1_transposed.dtypes)
    # print(df1_transposed.columns)
    # print(df1_transposed[df1_transposed['download_enable'].isin([False])])

    df_nowait = df1_transposed[df1_transposed['downlink_enable'].isin([False])]

    output.write(f"Regular Fragments (nowait)\n"
                 # f"send_time =\n{df_nowait['send_time']}"
                 f"count: {df_nowait['send_time'].count()}\n"
                 f"sum: {df_nowait['send_time'].sum(axis=0, skipna=True)}\n"
                 f"mean: {df_nowait['send_time'].mean(axis=0, skipna=True)}\n"
                 f"std: {df_nowait['send_time'].std(axis=0, skipna=True)}\n"
                 f"\n")

    sheet.cell(row=6, column=current_column).value = df_nowait['send_time'].count()
    sheet.cell(row=8, column=current_column).value = df_nowait['send_time'].sum(axis=0, skipna=True)
    sheet.cell(row=9, column=current_column).value = df_nowait['send_time'].mean(axis=0, skipna=True)
    sheet.cell(row=10, column=current_column).value = df_nowait['send_time'].std(axis=0, skipna=True) if not isnan(df_nowait['send_time'].std(axis=0, skipna=True)) else 0

    df_wait = df1_transposed[df1_transposed['downlink_enable'].isin([True])]

    # output.write(f"Regular Fragments (wait)\n"
    #              # f"data =\n{df_wait}\n"
    #              f"send_time =\n{df_wait['send_time']}"
    #              f"count: {df_wait['send_time'].count()}\n"
    #              f"sum: {df_wait['send_time'].sum(axis=0, skipna=True)}\n"
    #              f"mean: {df_wait['send_time'].mean(axis=0, skipna=True)}\n"
    #              f"std: {df_wait['send_time'].std(axis=0, skipna=True)}\n"
    #              f"\n")

    if len(df_wait[df_wait['RULE_ID'] == "000"]) != 0:
        df_all0 = df_wait[df_wait['FCN'].isin(['000'])]
        df_all1 = df_wait[df_wait['FCN'].isin(['111'])]
    else:
        df_all0 = df_wait[df_wait['FCN'].isin(['00000'])]
        df_all1 = df_wait[df_wait['FCN'].isin(['11111'])]

    output.write(f"Fragments - downlink requested - ALL 0\n"
                 f"data =\n{df_all0}\n"
                 f"send_time =\n{df_all0['send_time']}"
                 f"count: {df_all0['send_time'].count()}\n"
                 f"ul_errors: {df_all0[df_all0['ack_received'].isin([False])]['ack_received'].count()}\n"
                 f"all0_received: {df_all0[df_all0['ack_received'].isin([True])]['ack_received'].count()}\n"
                 f"sum: {df_all0['send_time'].sum(axis=0, skipna=True)}\n"
                 f"mean: {df_all0['send_time'].mean(axis=0, skipna=True)}\n"
                 f"std: {df_all0['send_time'].std(axis=0, skipna=True)}\n"
                 f"\n")

    sheet.cell(row=12, column=current_column).value = df_all0['send_time'].count()
    sheet.cell(row=13, column=current_column).value = df_all0[df_all0['ack_received'].isin([False])]['ack_received'].count()
    sheet.cell(row=15, column=current_column).value = df_all0[df_all0['ack_received'].isin([True])]['ack_received'].count()
    sheet.cell(row=16, column=current_column).value = df_all0['send_time'].sum(axis=0, skipna=True)
    sheet.cell(row=17, column=current_column).value = df_all0['send_time'].mean(axis=0, skipna=True) if not isnan(df_all0['send_time'].mean(axis=0, skipna=True)) else 0
    sheet.cell(row=18, column=current_column).value = df_all0['send_time'].std(axis=0, skipna=True) if not isnan(df_all0['send_time'].std(axis=0, skipna=True)) else 0

    output.write(f"Fragments - downlink requested - ALL 1\n"
                 f"data =\n{df_all1}\n"
                 f"send_time =\n{df_all1['send_time']}"
                 f"count: {df_all1['send_time'].count()}\n"
                 f"ul_errors: {df_all1[df_all1['ack_received'].isin([False])]['ack_received'].count()}\n"
                 f"all1_received: {df_all1[df_all1['ack_received'].isin([True])]['ack_received'].count()}\n"
                 f"sum: {df_all1['send_time'].sum(axis=0, skipna=True)}\n"
                 f"mean: {df_all1['send_time'].mean(axis=0, skipna=True)}\n"
                 f"std: {df_all1['send_time'].std(axis=0, skipna=True)}\n"
                 f"\n")

    sheet.cell(row=20, column=current_column).value = df_all1['send_time'].count()
    sheet.cell(row=21, column=current_column).value = df_all1[df_all1['ack_received'].isin([False])]['ack_received'].count()
    sheet.cell(row=23, column=current_column).value = df_all1[df_all1['ack_received'].isin([True])]['ack_received'].count()
    sheet.cell(row=24, column=current_column).value = df_all1['send_time'].sum(axis=0, skipna=True)
    sheet.cell(row=25, column=current_column).value = df_all1['send_time'].mean(axis=0, skipna=True)
    sheet.cell(row=26, column=current_column).value = df_all1['send_time'].std(axis=0, skipna=True) if not isnan(df_all1['send_time'].std(axis=0, skipna=True)) else 0

    # df1_transposed.to_excel('test_stats_2.2.xlsx', engine='xlsxwriter')
    output.write(f"Transmission Time (excluding code overhead): {df1_transposed['send_time'].sum(axis=0, skipna=True)}\n\n")
    sheet.cell(row=27, column=current_column).value = df1_transposed['send_time'].sum(axis=0, skipna=True)


if __name__ == '__main__':

    os.chdir("stats")
    with open("output.txt", 'w') as output_file:
        for foldername in ["0_ul", "10_ul", "20_ul", "10_uldl", "20_uldl"]:
            if foldername in ["0_ul", "10_ul", "20_ul"]:
                spreadsheet_name = 'Template Results UL.xlsx'
            else:
                spreadsheet_name = 'Template Results ULDL.xlsx'

            spreadsheet = openpyxl.load_workbook(filename=spreadsheet_name)
            print(f"Current folder: {foldername}")
            output_file.write(f"============== FOLDER: {foldername} ==============\n\n")
            for filename in os.listdir(f'results/{foldername}'):
                print(f"Extracting data from {foldername}/{filename}")
                extract_data(f"results/{foldername}/{filename}", output_file)
            spreadsheet.save(spreadsheet_name)
